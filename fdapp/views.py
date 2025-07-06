from django.shortcuts import render, redirect, get_object_or_404
from .models import FixedDeposit
from datetime import datetime, timedelta, date
import openpyxl
from django.http import HttpResponse
from .forms import FixedDepositForm
from decimal import Decimal
from datetime import datetime, date, timedelta
import openpyxl
from dateutil.relativedelta import relativedelta
from django.views.decorators.csrf import csrf_exempt



def home(request):
    fds = FixedDeposit.objects.all()

    # Unique financial years (calculated from start_date)
    fy_set = set()
    for fd in fds:
        fy = get_financial_year(fd.start_date)
        fy_set.add(fy)
    fy_list = sorted(fy_set, reverse=True)

    selected_fy = request.GET.get('fy')

    if selected_fy:
        fy_start = int(selected_fy.split('-')[0])
        fy_start_date = date(fy_start, 4, 1)
        fy_end_date = date(fy_start + 1, 3, 31)
        fds = fds.filter(start_date__gte=fy_start_date, start_date__lte=fy_end_date)

    return render(request, 'home.html', {
        'fds': fds,
        'fy_list': fy_list,
        'selected_fy': selected_fy
    })




def fd_list(request):
    fds = FixedDeposit.objects.all()
    return render(request, 'fd_list.html', {'fds': fds})


    
    
def get_quarter_dates(financial_year, quarter):
    try:
        start_year = int(financial_year.split('-')[0])
    except:
        return None, None

    if quarter == 'Q1':
        return date(start_year, 4, 1), date(start_year, 6, 30)
    elif quarter == 'Q2':
        return date(start_year, 7, 1), date(start_year, 9, 30)
    elif quarter == 'Q3':
        return date(start_year, 10, 1), date(start_year, 12, 31)
    elif quarter == 'Q4':
        return date(start_year + 1, 1, 1), date(start_year + 1, 3, 31)
    else:
        return None, None



def preview_uploaded_data(request):
    data = request.session.get('uploaded_data', [])
    filename = request.session.get('filename', 'Uploaded File')
    return render(request, 'preview.html', {'data': data, 'filename': filename})


def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        preview_data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 8:
                continue

            customer_id, fd_number, bank_name, principal, start_date, maturity_date, accumulation, rate = row

            try:
                if isinstance(rate, str) and "%" in rate:
                    rate_decimal = float(rate.replace('%', '').strip()) / 100
                else:
                    rate_decimal = float(rate)
                rate_percent = round(rate_decimal * 100, 2)
                principal = int(principal)
            except:
                continue

            try:
                if not isinstance(start_date, date):
                    start_date = datetime.strptime(str(start_date), '%Y-%m-%d').date()
                if not isinstance(maturity_date, date):
                    maturity_date = datetime.strptime(str(maturity_date), '%Y-%m-%d').date()
            except:
                continue

            accumulation_flag = "Yes" if str(accumulation).lower() == "yes" else "No"
            interest_type = "Actual" if accumulation_flag == "Yes" else "Accrued"
            financial_year = get_financial_year(start_date)

            balance = principal
            total_interest = 0
            total_tds = 0
            current_date = start_date

            while current_date < maturity_date:
                end_date = current_date + timedelta(days=91)
                if end_date > maturity_date:
                    end_date = maturity_date

                days = (end_date - current_date).days
                interest = (balance * rate_decimal * days) / (100 * 365)

                if accumulation_flag == "Yes" and days >= 91:
                    tds = interest * 0.10
                    net_interest = interest - tds
                    balance += net_interest
                else:
                    tds = 0
                    net_interest = 0

                total_interest += interest
                total_tds += tds
                current_date = end_date

            closing_balance = round(balance, 2)

            # ✅ Save to DB
            FixedDeposit.objects.create(
                customer_id=customer_id,
                fd_number=fd_number,
                bank_name=bank_name,
                principal=principal,
                rate=rate_decimal * 100,  # store as percent
                start_date=start_date,
                maturity_date=maturity_date,
                financial_year=financial_year,
                interest=round(total_interest, 2),
                tds=round(total_tds, 2),
                closing_balance=closing_balance,
                interest_type=interest_type,
                accumulation=accumulation_flag
            )

            # ✅ Preview table data
            preview_data.append([
                customer_id,
                fd_number,
                bank_name,
                f"₹{principal}",
                f"{rate_percent}%",
                start_date.strftime('%Y-%m-%d'),
                maturity_date.strftime('%Y-%m-%d'),
                accumulation_flag,
                interest_type,
                f"₹{round(total_interest, 2)}",
                f"₹{round(total_tds, 2)}",
                f"₹{closing_balance}"
            ])

        request.session['uploaded_data'] = preview_data
        request.session['filename'] = excel_file.name
        return redirect('preview_uploaded_data')

    return render(request, 'upload_excel.html')




def get_next_quarter_end(start_date, cutoff_date):
    quarters = [
        (4, 1), (7, 1), (10, 1), (1, 1)  # Apr 1, Jul 1, Oct 1, Jan 1
    ]
    year = start_date.year
    for m, d in quarters:
        q_start = date(year, m, d)
        if start_date < q_start:
            return min(q_start, cutoff_date)
    return min(date(year + 1, 4, 1), cutoff_date)



def get_financial_year(date_obj):
    if date_obj.month >= 4:
        return f"{date_obj.year}-{date_obj.year + 1}"
    else:
        return f"{date_obj.year - 1}-{date_obj.year}"


#====================================add fd=======================================


def add_fd(request):
    if request.method == 'POST':
        form = FixedDepositForm(request.POST)
        if form.is_valid():
            fd = form.save(commit=False)
            fd.principal = int(form.cleaned_data['principal'])
            fd.save()
            return redirect('/')
    else:
        form = FixedDepositForm()
    return render(request, 'add_fd.html', {'form': form})

#=================================================================================================



#==========================================Report code=================================
def generate_report(request):
    from .models import FixedDeposit
    from decimal import Decimal
    from datetime import datetime, date, timedelta
    from dateutil.relativedelta import relativedelta

    fds = FixedDeposit.objects.all().order_by('start_date')
    report = []
    cutoff_date = None
    selected_quarter = None
    today = date.today()

    if request.method == 'POST':
        cutoff_str = request.POST.get('cutoff_date')
        selected_quarter = request.POST.get('quarter')
        
        if not cutoff_str or not selected_quarter:
            return render(request, 'report.html', {
                'fds': fds,
                'report': [],
                'cutoff_date': None,
                'selected_quarter': selected_quarter,
                'error': "Please select both cutoff date and quarter."
            })

        try:
            cutoff_date = datetime.strptime(cutoff_str, '%Y-%m-%d').date()
        except ValueError:
            return render(request, 'report.html', {
                'fds': fds,
                'report': [],
                'cutoff_date': None,
                'selected_quarter': selected_quarter,
                'error': "Invalid date format. Please use YYYY-MM-DD."
            })

        # Determine financial year (April to March)
        fy = cutoff_date.year if cutoff_date.month >= 4 else cutoff_date.year - 1
        quarter_ranges = {
            'Q1': (date(fy, 4, 1), date(fy, 6, 30)),
            'Q2': (date(fy, 7, 1), date(fy, 9, 30)),
            'Q3': (date(fy, 10, 1), date(fy, 12, 31)),
            'Q4': (date(fy + 1, 1, 1), date(fy + 1, 3, 31)),
        }

        quarter_start, quarter_end = quarter_ranges[selected_quarter]
        if cutoff_date < quarter_end:
            quarter_end = cutoff_date

        for fd in fds:
            # Skip FDs that don't fall in the reporting period
            if fd.start_date > quarter_end or fd.maturity_date < quarter_start:
                continue

            balance = Decimal(fd.principal)
            rate = Decimal(fd.rate) / Decimal(100)
            rows = []
            total_interest = Decimal('0.00')
            total_tds = Decimal('0.00')
            current_balance = balance

            # Calculate the 3-month period from start date
            acc_start = fd.start_date
            acc_end_3month = acc_start + relativedelta(months=3)
            
            # Check if full 3-month period is within this quarter
            full_3month_in_quarter = (acc_end_3month <= quarter_end)
            
            # Determine the actual end date for the period
            acc_end = min(acc_end_3month if full_3month_in_quarter else quarter_end, 
                         fd.maturity_date, 
                         cutoff_date)

            if acc_start < acc_end:
                acc_days = (acc_end - acc_start).days
                acc_interest = (balance * rate * acc_days) / Decimal(365)
                
                if fd.accumulation == 'Yes' and full_3month_in_quarter:
                    # Full 3-month accumulation within quarter
                    int_type = 'Accumulated'
                    tds = acc_interest * Decimal('0.10')
                    net_interest = acc_interest - tds
                    current_balance = balance + net_interest
                else:
                    # Either not accumulation or doesn't complete 3 months in quarter
                    int_type = 'Accrued'
                    tds = Decimal('0.00')
                    net_interest = acc_interest
                    current_balance = balance

                rows.append({
                    'quarter': selected_quarter,
                    'from': acc_start,
                    'to': acc_end,
                    'days': acc_days,
                    'type': int_type,
                    'interest': round(acc_interest, 2),
                    'tds': round(tds, 2),
                    'net_interest': round(net_interest, 2),
                    'opening_balance': round(balance, 2),
                    'closing_balance': round(current_balance, 2),
                })

                total_interest += acc_interest
                total_tds += tds

                # For cases where we have remaining days after 3-month period
                if full_3month_in_quarter:
                    remaining_start = acc_end + timedelta(days=1)
                    remaining_end = min(fd.maturity_date, quarter_end, cutoff_date)
                    
                    if remaining_start < remaining_end:
                        remaining_days = (remaining_end - remaining_start).days
                        remaining_interest = (current_balance * rate * remaining_days) / Decimal(365)
                        
                        rows.append({
                            'quarter': selected_quarter,
                            'from': remaining_start,
                            'to': remaining_end,
                            'days': remaining_days,
                            'type': 'Accrued',
                            'interest': round(remaining_interest, 2),
                            'tds': Decimal('0.00'),
                            'net_interest': round(remaining_interest, 2),
                            'opening_balance': round(current_balance, 2),
                            'closing_balance': round(current_balance, 2),
                        })

                        total_interest += remaining_interest

            report.append({
                'fd': fd,
                'rows': rows,
                'total_interest': round(total_interest, 2),
                'total_tds': round(total_tds, 2),
                'closing_balance': round(current_balance, 2) if rows else round(balance, 2),
                'interest_type': int_type if rows else 'Accrued',
                'rate_percent': f"{fd.rate:.2f}%",
                'is_current': fd.start_date <= today <= fd.maturity_date
            })

    return render(request, 'report.html', {
        'fds': fds,
        'report': report,
        'cutoff_date': cutoff_date.strftime('%Y-%m-%d') if cutoff_date else None,
        'selected_quarter': selected_quarter,
        'today': today
    })

#=============================================================================================================



@csrf_exempt
def export_to_excel(request):
    if request.method == "POST":
        quarter = request.POST.get('quarter')
        cutoff_date = request.POST.get('cutoff_date')

        if not quarter or not cutoff_date:
            return HttpResponse("Quarter or cutoff date not specified", status=400)

        try:
            cutoff_date = datetime.strptime(cutoff_date, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Invalid cutoff date format", status=400)

        fds = FixedDeposit.objects.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FD Quarter Report"

        headers = [
            "Customer ID", "FD Number", "Bank Name", "Principal", "Rate (%)",
            "Start Date", "Maturity Date", "Quarter", "From", "To", "Days",
            "Interest Type", "Interest", "TDS", "Net Interest", "Closing Balance"
        ]
        ws.append(headers)

        for fd in fds:
            fy = get_financial_year(fd.start_date)
            q_start, q_end = get_quarter_dates(fy, quarter)
            if not q_start or not q_end:
                continue

            start = max(fd.start_date, q_start)
            end = min(fd.maturity_date, q_end, cutoff_date)

            if start >= end:
                continue

            balance = Decimal(fd.principal)
            rate = Decimal(fd.rate) / Decimal(100)
            days = Decimal((end - start).days)
            interest = (balance * rate * days) / Decimal(365)

            if fd.accumulation == 'Yes' and days >= 91:
                tds = interest * Decimal('0.10')
                net_interest = interest - tds
                balance += net_interest
                interest_type = "Accumulated"
            else:
                tds = Decimal('0.00')
                net_interest = interest
                interest_type = "Accrued"

            closing_balance = round(balance, 2)

            ws.append([
                fd.customer_id,
                fd.fd_number,
                fd.bank_name,
                float(fd.principal),
                f"{round(float(fd.rate), 2)}%",
                fd.start_date.strftime('%Y-%m-%d'),
                fd.maturity_date.strftime('%Y-%m-%d'),
                quarter,
                start.strftime('%Y-%m-%d'),
                end.strftime('%Y-%m-%d'),
                int(days),
                interest_type,
                float(round(interest, 2)),
                float(round(tds, 2)),
                float(round(net_interest, 2)),
                float(closing_balance)
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=fd_report_{quarter}.xlsx'
        wb.save(response)
        return response

    return HttpResponse("Invalid request method", status=405)






#=========================Edit & Delete===========================================
def edit_fd(request, pk):
    fd = get_object_or_404(FixedDeposit, pk=pk)
    if request.method == 'POST':
        form = FixedDepositForm(request.POST, instance=fd)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = FixedDepositForm(instance=fd)
    return render(request, 'edit_fd.html', {'form': form, 'fd': fd})

# Delete FD View
def delete_fd(request, pk):
    fd = get_object_or_404(FixedDeposit, pk=pk)
    if request.method == 'POST':
        fd.delete()
        return redirect('home')
    return render(request, 'confirm_delete.html', {'fd': fd})